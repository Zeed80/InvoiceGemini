#!/usr/bin/env python3
"""
Упрощенный дизайнер шаблонов как fallback вариант.
Используется если основной ExportTemplateDesigner не работает.
"""

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QTextEdit, QMessageBox, QTabWidget, QWidget,
    QFormLayout, QLineEdit, QComboBox, QCheckBox
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont

class SimpleTemplateDesigner(QDialog):
    """
    Упрощенный дизайнер шаблонов экспорта.
    Меньше функций, но более стабильный.
    """
    
    # Сигналы
    template_saved = pyqtSignal(str)
    template_applied = pyqtSignal(dict)
    
    def __init__(self, current_results=None, parent=None):
        super().__init__(parent)
        self.current_results = current_results or {}
        
        self.init_ui()
        self.populate_preview()
    
    def init_ui(self):
        """Инициализация упрощенного интерфейса"""
        self.setWindowTitle("🎨 Простой дизайнер шаблонов - InvoiceGemini")
        self.setMinimumSize(800, 600)
        self.resize(1000, 700)
        
        layout = QVBoxLayout(self)
        
        # Заголовок
        title_label = QLabel("Упрощенный дизайнер шаблонов экспорта")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Вкладки
        tabs = QTabWidget()
        
        # Вкладка настроек
        settings_tab = self.create_settings_tab()
        tabs.addTab(settings_tab, "⚙️ Настройки")
        
        # Вкладка предпросмотра
        preview_tab = self.create_preview_tab()
        tabs.addTab(preview_tab, "👁️ Предпросмотр")
        
        layout.addWidget(tabs)
        
        # Кнопки
        button_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("💾 Сохранить шаблон")
        self.save_btn.clicked.connect(self.save_template)
        
        self.export_btn = QPushButton("📤 Экспорт")
        self.export_btn.clicked.connect(self.export_data)
        
        self.close_btn = QPushButton("❌ Закрыть")
        self.close_btn.clicked.connect(self.close)
        
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)
        
        layout.addLayout(button_layout)
    
    def create_settings_tab(self):
        """Создает вкладку с настройками"""
        widget = QWidget()
        layout = QFormLayout(widget)
        
        # Основные настройки
        self.template_name = QLineEdit("Новый шаблон")
        layout.addRow("📝 Название шаблона:", self.template_name)
        
        self.template_format = QComboBox()
        self.template_format.addItems(["Excel (.xlsx)", "CSV (.csv)", "HTML (.html)"])
        layout.addRow("📄 Формат экспорта:", self.template_format)
        
        # Настройки содержимого
        self.include_header = QCheckBox("Включить заголовок")
        self.include_header.setChecked(True)
        layout.addRow("📋 Заголовок:", self.include_header)
        
        self.header_text = QLineEdit("Отчёт по обработке счетов")
        layout.addRow("   Текст заголовка:", self.header_text)
        
        self.include_footer = QCheckBox("Включить подвал")
        self.include_footer.setChecked(True)
        layout.addRow("📋 Подвал:", self.include_footer)
        
        self.footer_text = QLineEdit("Создано InvoiceGemini")
        layout.addRow("   Текст подвала:", self.footer_text)
        
        self.include_timestamp = QCheckBox("Включить время создания")
        self.include_timestamp.setChecked(True)
        layout.addRow("🕒 Время создания:", self.include_timestamp)
        
        return widget
    
    def create_preview_tab(self):
        """Создает вкладку предпросмотра"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        preview_label = QLabel("Предварительный просмотр:")
        layout.addWidget(preview_label)
        
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        layout.addWidget(self.preview_text)
        
        refresh_btn = QPushButton("🔄 Обновить предпросмотр")
        refresh_btn.clicked.connect(self.populate_preview)
        layout.addWidget(refresh_btn)
        
        return widget
    
    def populate_preview(self):
        """Заполняет предпросмотр данными"""
        preview_html = """
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .header { text-align: center; font-size: 18px; font-weight: bold; margin-bottom: 20px; }
                .footer { text-align: center; font-size: 12px; color: gray; margin-top: 20px; }
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
        """
        
        # Заголовок
        if hasattr(self, 'include_header') and self.include_header.isChecked():
            header_text = self.header_text.text() if hasattr(self, 'header_text') else "Отчёт по обработке счетов"
            preview_html += f'<div class="header">{header_text}</div>\n'
        
        # Таблица данных
        if self.current_results:
            preview_html += "<table>\n"
            
            if isinstance(self.current_results, dict) and "batch_results" in self.current_results:
                # Пакетная обработка
                batch_data = self.current_results["batch_results"]
                if batch_data:
                    # Заголовки
                    headers = list(batch_data[0].keys())
                    preview_html += "<tr>"
                    for header in headers:
                        preview_html += f"<th>{header}</th>"
                    preview_html += "</tr>\n"
                    
                    # Данные
                    for row_data in batch_data:
                        preview_html += "<tr>"
                        for header in headers:
                            value = row_data.get(header, "")
                            # Обрабатываем многострочный текст
                            if "\n" in str(value):
                                value = str(value).replace("\n", "<br/>")
                            preview_html += f"<td>{value}</td>"
                        preview_html += "</tr>\n"
            else:
                # Одиночная обработка
                preview_html += "<tr><th>Поле</th><th>Значение</th></tr>\n"
                for key, value in self.current_results.items():
                    # Обрабатываем многострочный текст
                    if "\n" in str(value):
                        value = str(value).replace("\n", "<br/>")
                    preview_html += f"<tr><td>{key}</td><td>{value}</td></tr>\n"
            
            preview_html += "</table>\n"
        else:
            preview_html += "<p><em>Нет данных для отображения</em></p>\n"
        
        # Временная метка
        if hasattr(self, 'include_timestamp') and self.include_timestamp.isChecked():
            from datetime import datetime
            timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            preview_html += f'<p style="font-size: 12px; color: gray;">Создано: {timestamp}</p>\n'
        
        # Подвал
        if hasattr(self, 'include_footer') and self.include_footer.isChecked():
            footer_text = self.footer_text.text() if hasattr(self, 'footer_text') else "Создано InvoiceGemini"
            preview_html += f'<div class="footer">{footer_text}</div>\n'
        
        preview_html += "</body></html>"
        
        self.preview_text.setHtml(preview_html)
    
    def save_template(self):
        """Сохраняет шаблон"""
        template_name = self.template_name.text()
        
        template_data = {
            "name": template_name,
            "format": self.template_format.currentText(),
            "include_header": self.include_header.isChecked(),
            "header_text": self.header_text.text(),
            "include_footer": self.include_footer.isChecked(),
            "footer_text": self.footer_text.text(),
            "include_timestamp": self.include_timestamp.isChecked()
        }
        
        QMessageBox.information(
            self,
            "Шаблон сохранен",
            f"Шаблон '{template_name}' успешно сохранен!\n\nНастройки:\n"
            f"• Формат: {template_data['format']}\n"
            f"• Заголовок: {'Да' if template_data['include_header'] else 'Нет'}\n"
            f"• Подвал: {'Да' if template_data['include_footer'] else 'Нет'}\n"
            f"• Время создания: {'Да' if template_data['include_timestamp'] else 'Нет'}"
        )
        
        self.template_saved.emit(template_name)
    
    def export_data(self):
        """Экспортирует данные с текущими настройками"""
        if not self.current_results:
            QMessageBox.warning(
                self,
                "Нет данных",
                "Нет данных для экспорта. Сначала обработайте документы."
            )
            return
        
        from PyQt6.QtWidgets import QFileDialog
        
        format_text = self.template_format.currentText()
        if "Excel" in format_text:
            filter_text = "Excel files (*.xlsx)"
            default_ext = ".xlsx"
        elif "CSV" in format_text:
            filter_text = "CSV files (*.csv)"
            default_ext = ".csv"
        else:
            filter_text = "HTML files (*.html)"
            default_ext = ".html"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт данных",
            f"export{default_ext}",
            filter_text
        )
        
        if file_path:
            try:
                if file_path.endswith('.html'):
                    self.export_to_html(file_path)
                elif file_path.endswith('.csv'):
                    self.export_to_csv(file_path)
                elif file_path.endswith('.xlsx'):
                    self.export_to_excel(file_path)
                
                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Данные успешно экспортированы в:\n{file_path}"
                )
                
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка экспорта",
                    f"Не удалось экспортировать данные:\n{str(e)}"
                )
    
    def export_to_html(self, file_path):
        """Экспорт в HTML"""
        html_content = self.preview_text.toHtml()
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def export_to_csv(self, file_path):
        """Экспорт в CSV"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            if isinstance(self.current_results, dict) and "batch_results" in self.current_results:
                # Пакетная обработка
                batch_data = self.current_results["batch_results"]
                if batch_data:
                    headers = list(batch_data[0].keys())
                    writer.writerow(headers)
                    for row_data in batch_data:
                        row = [row_data.get(h, "") for h in headers]
                        writer.writerow(row)
            else:
                # Одиночная обработка
                writer.writerow(["Поле", "Значение"])
                for key, value in self.current_results.items():
                    writer.writerow([key, value])
    
    def export_to_excel(self, file_path):
        """Экспорт в Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Результаты обработки"
            
            row_num = 1
            
            # Заголовок
            if self.include_header.isChecked():
                ws.merge_cells(f'A{row_num}:B{row_num}')
                header_cell = ws[f'A{row_num}']
                header_cell.value = self.header_text.text()
                header_cell.font = Font(size=16, bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                row_num += 2
            
            # Данные
            if isinstance(self.current_results, dict) and "batch_results" in self.current_results:
                # Пакетная обработка
                batch_data = self.current_results["batch_results"]
                if batch_data:
                    headers = list(batch_data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row_num, column=col, value=header).font = Font(bold=True)
                    row_num += 1
                    
                    for row_data in batch_data:
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row_num, column=col, value=row_data.get(header, ""))
                        row_num += 1
            else:
                # Одиночная обработка
                ws.cell(row=row_num, column=1, value="Поле").font = Font(bold=True)
                ws.cell(row=row_num, column=2, value="Значение").font = Font(bold=True)
                row_num += 1
                
                for key, value in self.current_results.items():
                    ws.cell(row=row_num, column=1, value=key)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1
            
            # Подвал
            if self.include_footer.isChecked():
                row_num += 1
                ws.merge_cells(f'A{row_num}:B{row_num}')
                footer_cell = ws[f'A{row_num}']
                footer_cell.value = self.footer_text.text()
                footer_cell.font = Font(size=10, italic=True)
                footer_cell.alignment = Alignment(horizontal='center')
            
            # Временная метка
            if self.include_timestamp.isChecked():
                row_num += 1
                from datetime import datetime
                timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                ws.merge_cells(f'A{row_num}:B{row_num}')
                timestamp_cell = ws[f'A{row_num}']
                timestamp_cell.value = f"Создано: {timestamp}"
                timestamp_cell.font = Font(size=9)
                timestamp_cell.alignment = Alignment(horizontal='center')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError, ValueError) as e:
                        # Ячейка может содержать None или несовместимый тип
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except ImportError:
            # Fallback без openpyxl
            raise Exception("Модуль openpyxl не установлен. Используйте CSV или HTML формат.")
    
    def closeEvent(self, event):
        """Обработка закрытия диалога"""
        event.accept() 