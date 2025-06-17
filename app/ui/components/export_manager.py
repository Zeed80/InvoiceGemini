"""
Менеджер экспорта данных в различные форматы.
"""

import os
import csv
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportManager:
    """Менеджер для экспорта данных в различные форматы"""
    
    SUPPORTED_FORMATS = {
        'csv': 'CSV файлы (*.csv)',
        'json': 'JSON файлы (*.json)',
        'xml': 'XML файлы (*.xml)',
        'excel': 'Excel файлы (*.xlsx)',
        'html': 'HTML файлы (*.html)',
        'pdf': 'PDF файлы (*.pdf)'
    }
    
    def __init__(self):
        """Инициализация менеджера экспорта"""
        self.encoding = 'utf-8'
        
    def export_data(self, data: List[Dict[str, Any]], file_path: str, 
                   format_type: str, options: Optional[Dict] = None) -> bool:
        """
        Экспорт данных в указанный формат.
        
        Args:
            data: Список словарей с данными
            file_path: Путь для сохранения файла
            format_type: Тип формата (csv, json, xml, excel, html, pdf)
            options: Дополнительные опции экспорта
            
        Returns:
            bool: True если экспорт успешен
        """
        try:
            options = options or {}
            
            if format_type == 'csv':
                return self._export_csv(data, file_path, options)
            elif format_type == 'json':
                return self._export_json(data, file_path, options)
            elif format_type == 'xml':
                return self._export_xml(data, file_path, options)
            elif format_type == 'excel':
                return self._export_excel(data, file_path, options)
            elif format_type == 'html':
                return self._export_html(data, file_path, options)
            elif format_type == 'pdf':
                return self._export_pdf(data, file_path, options)
            else:
                logger.error(f"Неподдерживаемый формат: {format_type}")
                return False
                
        except Exception as e:
            logger.error(f"Ошибка экспорта в {format_type}: {e}")
            return False
            
    def _export_csv(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в CSV формат"""
        try:
            # Получаем все уникальные поля
            all_fields = set()
            for row in data:
                all_fields.update(row.keys())
            
            fields = sorted(list(all_fields))
            
            # Параметры экспорта
            delimiter = options.get('delimiter', ';')
            include_header = options.get('include_header', True)
            
            with open(file_path, 'w', newline='', encoding=self.encoding) as f:
                writer = csv.DictWriter(f, fieldnames=fields, delimiter=delimiter)
                
                if include_header:
                    # Можем использовать русские названия полей
                    header_mapping = options.get('header_mapping', {})
                    if header_mapping:
                        writer.writerow({field: header_mapping.get(field, field) 
                                       for field in fields})
                    else:
                        writer.writeheader()
                        
                writer.writerows(data)
                
            logger.info(f"Данные экспортированы в CSV: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта CSV: {e}")
            return False
            
    def _export_json(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в JSON формат"""
        try:
            # Параметры форматирования
            indent = options.get('indent', 2)
            ensure_ascii = options.get('ensure_ascii', False)
            
            # Добавляем метаданные
            export_data = {
                'export_date': datetime.now().isoformat(),
                'total_records': len(data),
                'format_version': '1.0',
                'data': data
            }
            
            if options.get('include_metadata', True):
                with open(file_path, 'w', encoding=self.encoding) as f:
                    json.dump(export_data, f, indent=indent, 
                             ensure_ascii=ensure_ascii)
            else:
                with open(file_path, 'w', encoding=self.encoding) as f:
                    json.dump(data, f, indent=indent, 
                             ensure_ascii=ensure_ascii)
                    
            logger.info(f"Данные экспортированы в JSON: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта JSON: {e}")
            return False
            
    def _export_xml(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в XML формат"""
        try:
            # Создаем корневой элемент
            root = ET.Element('invoices')
            root.set('version', '1.0')
            root.set('export_date', datetime.now().isoformat())
            
            # Добавляем записи
            for idx, record in enumerate(data):
                invoice_elem = ET.SubElement(root, 'invoice')
                invoice_elem.set('id', str(idx + 1))
                
                for key, value in record.items():
                    if value is not None:
                        field_elem = ET.SubElement(invoice_elem, self._sanitize_xml_tag(key))
                        field_elem.text = str(value)
                        
            # Форматируем XML
            xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
            
            # Сохраняем
            with open(file_path, 'w', encoding=self.encoding) as f:
                f.write(xml_str)
                
            logger.info(f"Данные экспортированы в XML: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта XML: {e}")
            return False
            
    def _export_excel(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в Excel формат"""
        try:
            # Создаем DataFrame
            df = pd.DataFrame(data)
            
            # Переименовываем колонки если есть маппинг
            column_mapping = options.get('column_mapping', {})
            if column_mapping:
                df.rename(columns=column_mapping, inplace=True)
                
            # Создаем Excel файл с форматированием
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Данные счетов', index=False)
                
                # Получаем worksheet для форматирования
                worksheet = writer.sheets['Данные счетов']
                
                # Форматирование заголовков
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", 
                                        fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    
                # Автоматическая ширина колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                            
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                # Добавляем сводную информацию на отдельный лист
                if options.get('include_summary', True):
                    summary_df = self._create_summary(data)
                    summary_df.to_excel(writer, sheet_name='Сводка', index=False)
                    
            logger.info(f"Данные экспортированы в Excel: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта Excel: {e}")
            return False
            
    def _export_html(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в HTML формат"""
        try:
            df = pd.DataFrame(data)
            
            # Переименовываем колонки если есть маппинг
            column_mapping = options.get('column_mapping', {})
            if column_mapping:
                df.rename(columns=column_mapping, inplace=True)
                
            # Создаем HTML
            html_template = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Экспорт данных счетов</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        h1 {{
            color: #333;
            text-align: center;
        }}
        .info {{
            background-color: #e3f2fd;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 20px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            background-color: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #1976d2;
            color: white;
            padding: 12px;
            text-align: left;
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .footer {{
            margin-top: 20px;
            text-align: center;
            color: #666;
            font-size: 0.9em;
        }}
    </style>
</head>
<body>
    <h1>Данные счетов</h1>
    <div class="info">
        <p><strong>Дата экспорта:</strong> {export_date}</p>
        <p><strong>Всего записей:</strong> {total_records}</p>
    </div>
    {table}
    <div class="footer">
        <p>Экспортировано из InvoiceGemini</p>
    </div>
</body>
</html>
"""
            
            # Генерируем таблицу
            table_html = df.to_html(index=False, classes='invoice-table', 
                                  table_id='data-table', escape=False)
            
            # Подставляем данные в шаблон
            final_html = html_template.format(
                export_date=datetime.now().strftime('%d.%m.%Y %H:%M'),
                total_records=len(data),
                table=table_html
            )
            
            # Сохраняем
            with open(file_path, 'w', encoding=self.encoding) as f:
                f.write(final_html)
                
            logger.info(f"Данные экспортированы в HTML: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта HTML: {e}")
            return False
            
    def _export_pdf(self, data: List[Dict], file_path: str, options: Dict) -> bool:
        """Экспорт в PDF формат"""
        try:
            # Для PDF используем HTML как промежуточный формат
            # и конвертируем через wkhtmltopdf или аналог
            
            # Временно экспортируем в HTML
            import tempfile
            temp_html = tempfile.NamedTemporaryFile(mode='w', suffix='.html', 
                                                   delete=False, encoding=self.encoding)
            temp_html_path = temp_html.name
            temp_html.close()
            
            # Экспортируем в HTML
            if not self._export_html(data, temp_html_path, options):
                return False
                
            # Конвертируем HTML в PDF
            try:
                import pdfkit
                pdfkit.from_file(temp_html_path, file_path)
                logger.info(f"Данные экспортированы в PDF: {file_path}")
                return True
            except ImportError:
                logger.warning("pdfkit не установлен, используем альтернативный метод")
                
                # Альтернативный метод через weasyprint
                try:
                    from weasyprint import HTML
                    HTML(filename=temp_html_path).write_pdf(file_path)
                    logger.info(f"Данные экспортированы в PDF через weasyprint: {file_path}")
                    return True
                except ImportError:
                    logger.error("Ни pdfkit, ни weasyprint не установлены для экспорта в PDF")
                    return False
            finally:
                # Удаляем временный файл
                if os.path.exists(temp_html_path):
                    os.unlink(temp_html_path)
                    
        except Exception as e:
            logger.error(f"Ошибка экспорта PDF: {e}")
            return False
            
    def _sanitize_xml_tag(self, tag: str) -> str:
        """Очистка имени тега для XML"""
        # Заменяем недопустимые символы
        tag = tag.replace(' ', '_')
        tag = tag.replace('-', '_')
        tag = ''.join(c for c in tag if c.isalnum() or c == '_')
        
        # Убеждаемся, что тег не начинается с цифры
        if tag and tag[0].isdigit():
            tag = 'field_' + tag
            
        return tag or 'field'
        
    def _create_summary(self, data: List[Dict]) -> pd.DataFrame:
        """Создание сводной информации"""
        summary = {
            'Показатель': [],
            'Значение': []
        }
        
        # Общее количество
        summary['Показатель'].append('Всего счетов')
        summary['Значение'].append(len(data))
        
        # Попытка посчитать суммы
        try:
            total_amounts = []
            for record in data:
                for key, value in record.items():
                    if 'сумма' in key.lower() or 'total' in key.lower():
                        try:
                            amount = float(str(value).replace(',', '.').replace(' ', ''))
                            total_amounts.append(amount)
                        except:
                            pass
                            
            if total_amounts:
                summary['Показатель'].append('Общая сумма')
                summary['Значение'].append(f"{sum(total_amounts):,.2f}")
                
                summary['Показатель'].append('Средняя сумма')
                summary['Значение'].append(f"{sum(total_amounts)/len(total_amounts):,.2f}")
                
        except Exception as e:
            logger.warning(f"Не удалось подсчитать суммы: {e}")
            
        return pd.DataFrame(summary)
        
    def get_export_filters(self) -> Dict[str, str]:
        """Получение фильтров для диалога сохранения файла"""
        filters = []
        for format_type, description in self.SUPPORTED_FORMATS.items():
            filters.append(description)
            
        return ";;".join(filters) 